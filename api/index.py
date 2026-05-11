"""api/index.py — Rethink Food Route Generator (Flask/Vercel)"""

import csv
import io
import os
import sys
import uuid
import warnings
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from geopy.geocoders import Nominatim

# Project root is one level up from api/
sys.path.insert(0, str(Path(__file__).parent.parent))

from rethink_routes import (
    CACHE_FILE,
    DEPOT_BRONX_END,
    DEPOT_OTHER_END,
    DEPOT_START,
    MANIFEST_HEADERS,
    MAX_ROUTE_MILES,
    MAX_STOPS_SOFT,
    ROUTES,
    ZIP_OVERRIDES,
    build_map,
    check_stop_limit,
    clean_box,
    detect_household_clusters,
    geocode_stop,
    load_cache,
    optimize_route,
)

warnings.filterwarnings("ignore")

_ROOT = Path(__file__).parent.parent

app = Flask(
    __name__,
    template_folder=str(_ROOT / "templates"),
    static_folder=str(_ROOT / "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-in-prod")

# In-memory results store keyed by UUID (stored in session cookie)
_store: dict = {}


# ── Auth ──────────────────────────────────────────────────────────────────────

def _authed() -> bool:
    return bool(session.get("authed"))


# ── Pure helpers (ported from app.py) ─────────────────────────────────────────

def parse_excel(fileobj):
    wb = openpyxl.load_workbook(fileobj, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = rows[0]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    required = ["Member ID", "Box Size", "Address Line 1", "City", "State",
                "Zip", "Phone Number", "Delivery Instructions", "Status"]
    missing = [c for c in required if c not in col]
    if missing:
        return None, [f"Missing columns in spreadsheet: {missing}"]

    stops, flags = [], []
    for row in rows[1:]:
        row = tuple(row) + (None,) * max(0, len(headers) - len(row))
        status = str(row[col["Status"]] or "").strip().lower()
        if status != "active":
            continue

        zipcode = str(row[col["Zip"]] or "").replace(".0", "").strip().zfill(5)
        addr1   = str(row[col["Address Line 1"]] or "").strip()
        addr2   = str(row[col["Address Line 2"]] or "").strip() if "Address Line 2" in col else ""
        city    = str(row[col["City"]]   or "").strip()
        state   = str(row[col["State"]]  or "").strip()

        display_addr = addr1
        if addr2 and addr2.lower() not in ("none", ""):
            display_addr += f", {addr2}"
        display_addr += f", {city}, {state} {zipcode}"

        allergens = str(row[col["Meal Preferences/Allergens"]] or "").strip() \
            if "Meal Preferences/Allergens" in col else ""
        avail = str(row[col["Available Delivery Days"]] or "").strip() \
            if "Available Delivery Days" in col else ""
        notes_raw = " | ".join(filter(None, [
            str(row[col[c]] or "").strip()
            for c in ["Unnamed: 13", "Unnamed: 14", "Unnamed: 15"] if c in col
        ]))

        flag = ""
        if "cancel" in notes_raw.lower():
            flag = "Notes say 'cancel' but status is Active"
        elif "hold" in notes_raw.lower():
            flag = "Notes mention 'hold' — verify"
        if allergens and allergens.lower() not in ("none", ""):
            flag = (flag + " | " if flag else "") + f"Allergen: {allergens}"

        member_id = str(row[col["Member ID"]] or "").replace(".0", "")
        stops.append({
            "member_id":             member_id,
            "addr1":                 addr1,
            "addr2":                 addr2,
            "city":                  city,
            "state":                 state,
            "zipcode":               zipcode,
            "display_addr":          display_addr,
            "phone":                 str(row[col["Phone Number"]] or "").strip(),
            "box_size":              clean_box(row[col["Box Size"]]),
            "allergens":             allergens,
            "delivery_instructions": str(row[col["Delivery Instructions"]] or "").strip(),
            "available_days":        avail,
            "notes":                 notes_raw,
            "flag":                  flag,
            "latlon":                None,
        })
        if flag:
            flags.append(f"Member {member_id} ({display_addr}): {flag}")

    return stops, flags


def _build_summary_lines(route_info, contact, total_members):
    dep_end = DEPOT_BRONX_END if "Bronx" in route_info.get("depot_end", "") else DEPOT_OTHER_END
    bc = route_info.get("box_counts", {})
    box_parts = [f"{sz}: {bc[sz]}" for sz in ["Large", "Medium", "Small", "Four-Date"] if bc.get(sz)]
    return [
        f"Route: Route {route_info['letter']} — {route_info['name']} ({route_info['day']})",
        f"Date: {date.today().isoformat()}",
        f"Start: {DEPOT_START['label']} ({DEPOT_START['addr1']} {DEPOT_START['zipcode']})",
        f"End: {dep_end['label']} ({dep_end['addr1']} {dep_end['zipcode']})",
        f"Point of Contact: {contact}",
        f"Total Members: {total_members}",
        f"Boxes: {' | '.join(box_parts) if box_parts else '0'}",
    ]


def manifest_to_csv(ordered_stops, route_info=None, contact="Oscar"):
    buf = io.StringIO()
    w = csv.writer(buf)
    if route_info:
        for line in _build_summary_lines(route_info, contact, len(ordered_stops)):
            w.writerow([line])
        w.writerow(["---"])
    w.writerow(MANIFEST_HEADERS)
    for i, s in enumerate(ordered_stops, 1):
        w.writerow([
            i, s["member_id"], s["addr1"], s["addr2"], s["city"], s["zipcode"],
            s["phone"], s["box_size"], s["allergens"], s["delivery_instructions"],
            s["available_days"], s["notes"], s["flag"],
        ])
    return buf.getvalue().encode("utf-8")


def kitchen_to_csv(kitchen_rows):
    fields = ["Route", "Total Stops", "Large", "Medium", "Small", "Four-Date", "Unknown", "Allergen Notes"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fields)
    w.writeheader()
    w.writerows(kitchen_rows)
    if kitchen_rows:
        w.writerow({
            "Route":         "TOTAL",
            "Total Stops":   sum(r["Total Stops"] for r in kitchen_rows),
            "Large":         sum(r["Large"] for r in kitchen_rows),
            "Medium":        sum(r["Medium"] for r in kitchen_rows),
            "Small":         sum(r["Small"] for r in kitchen_rows),
            "Four-Date":     sum(r.get("Four-Date", 0) for r in kitchen_rows),
            "Unknown":       sum(r["Unknown"] for r in kitchen_rows),
            "Allergen Notes": "",
        })
    return buf.getvalue().encode("utf-8")


_XLSX_HEADERS = [
    "Stop", "Member ID", "Address", "Apt/Unit", "City", "Zip",
    "Phone", "Box Size", "Allergens", "Delivery Instructions",
    "Available Days", "Notes", "Flag", "Household",
]

_HOUSEHOLD_FILLS = [
    PatternFill(start_color="FFE8B0", end_color="FFE8B0", fill_type="solid"),
    PatternFill(start_color="B0D4FF", end_color="B0D4FF", fill_type="solid"),
    PatternFill(start_color="FFB0C8", end_color="FFB0C8", fill_type="solid"),
    PatternFill(start_color="B0FFB0", end_color="B0FFB0", fill_type="solid"),
    PatternFill(start_color="D4B0FF", end_color="D4B0FF", fill_type="solid"),
    PatternFill(start_color="FFD4B0", end_color="FFD4B0", fill_type="solid"),
]


def manifest_to_xlsx(ordered_stops, route_info, household_clusters, contact="Oscar"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"
    num_cols = len(_XLSX_HEADERS)

    summary = _build_summary_lines(route_info, contact, len(ordered_stops))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=summary[0].replace("Route: ", ""))
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="left")
    for r_idx, line in enumerate(summary[1:], start=2):
        ws.cell(row=r_idx, column=1, value=line)

    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    HEADER_ROW = 9
    for c_idx, hdr in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c_idx, value=hdr)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    red_font = Font(color="FF0000")
    group_letters = sorted(set(household_clusters.values()))
    group_fill_map = {
        letter: _HOUSEHOLD_FILLS[i % len(_HOUSEHOLD_FILLS)]
        for i, letter in enumerate(group_letters)
    }
    col_widths = [len(h) for h in _XLSX_HEADERS]

    for i, s in enumerate(ordered_stops):
        row_num = HEADER_ROW + 1 + i
        hh_letter = household_clusters.get(i, "")
        values = [
            i + 1, s["member_id"], s["addr1"], s["addr2"], s["city"],
            s["zipcode"], s["phone"], s["box_size"], s["allergens"],
            s["delivery_instructions"], s["available_days"], s["notes"],
            s["flag"], hh_letter,
        ]
        hh_fill = group_fill_map.get(hh_letter)
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            if s["flag"]:
                cell.font = red_font
            if hh_fill:
                cell.fill = hh_fill
            elif i % 2 == 1:
                cell.fill = alt_fill
            val_len = len(str(val)) if val else 0
            if val_len > col_widths[c_idx - 1]:
                col_widths[c_idx - 1] = val_len

    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = min(w + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run_generation(all_stops, all_flags, distance_cap=MAX_ROUTE_MILES, contact_name="Oscar"):
    """Geocode stops, optimize routes. Returns (results, kitchen_rows, flags)."""
    geocache   = load_cache()
    geolocator = Nominatim(user_agent="rethink_food_app")

    zip_to_routes = {}
    for letter, name, borough, day, zips in ROUTES:
        for z in zips:
            zip_to_routes.setdefault(z, []).append((letter, name, borough, day))

    route_buckets = {letter: [] for letter, *_ in ROUTES}
    flags = list(all_flags)

    for stop in all_stops:
        override_letter = ZIP_OVERRIDES.get(stop["zipcode"])
        if override_letter and override_letter in route_buckets:
            override_borough = next(
                (b for l, n, b, d, z in ROUTES if l == override_letter), "Queens"
            )
            route_buckets[override_letter].append((stop, override_borough))
            continue

        matching = zip_to_routes.get(stop["zipcode"], [])
        if not matching:
            flags.append(
                f"Member {stop['member_id']} (zip {stop['zipcode']}) "
                "not assigned to any route — add zip to ROUTES in rethink_routes.py"
            )
        elif len(matching) == 1:
            route_buckets[matching[0][0]].append((stop, matching[0][2]))
        else:
            best = min(matching, key=lambda r: len(route_buckets[r[0]]))
            route_buckets[best[0]].append((stop, best[2]))

    depot_start_latlon = geocode_stop(
        geolocator, DEPOT_START["addr1"], DEPOT_START["zipcode"], DEPOT_START["borough"], geocache)
    depot_bronx_latlon = geocode_stop(
        geolocator, DEPOT_BRONX_END["addr1"], DEPOT_BRONX_END["zipcode"], DEPOT_BRONX_END["borough"], geocache)
    depot_other_latlon = geocode_stop(
        geolocator, DEPOT_OTHER_END["addr1"], DEPOT_OTHER_END["zipcode"], DEPOT_OTHER_END["borough"], geocache)

    if not depot_start_latlon or not depot_bronx_latlon or not depot_other_latlon:
        raise RuntimeError(
            "Failed to geocode depot addresses. Check internet connection and try again."
        )

    for stop in all_stops:
        matching = zip_to_routes.get(stop["zipcode"])
        borough  = matching[0][2] if matching else "Queens"
        stop["latlon"] = geocode_stop(
            geolocator, stop["addr1"], stop["zipcode"], borough, geocache
        )
        if stop["latlon"] is None:
            flags.append(
                f"Member {stop['member_id']} ({stop['addr1']}, {stop['zipcode']}): "
                "geocoding failed — stop excluded from map"
            )

    results, kitchen_rows = [], []

    for letter, name, borough, day, zips in ROUTES:
        entries = route_buckets[letter]
        if not entries:
            continue
        geocoded = [e[0] for e in entries if e[0]["latlon"]]
        if not geocoded:
            continue

        depot_end_latlon = depot_bronx_latlon if borough == "Bronx" else depot_other_latlon
        depot_end_info   = DEPOT_BRONX_END    if borough == "Bronx" else DEPOT_OTHER_END

        ordered, orig_dist, opt_dist = optimize_route(
            geocoded, depot_start_latlon, depot_end_latlon
        )

        limit_warning = check_stop_limit(ordered)
        if limit_warning:
            flags.append(f"Route {letter} ({name.replace('_', ' ')}): {limit_warning}")

        distance_warning = None
        if opt_dist > distance_cap:
            distance_warning = f"{opt_dist:.1f} miles exceeds {distance_cap:.1f} mile cap"
            flags.append(f"Route {letter} ({name.replace('_', ' ')}): {distance_warning}")

        box_counts     = {"Large": 0, "Medium": 0, "Small": 0, "Four-Date": 0, "Unknown": 0}
        allergen_notes = []
        for s in ordered:
            box_counts[s["box_size"]] = box_counts.get(s["box_size"], 0) + 1
            if s["allergens"] and s["allergens"].lower() not in ("none", ""):
                allergen_notes.append(f"Member {s['member_id']}: {s['allergens']}")

        if sum(box_counts.values()) != len(ordered):
            flags.append(
                f"Route {letter} ({name.replace('_', ' ')}): box count mismatch — "
                f"{sum(box_counts.values())} boxes vs {len(ordered)} stops"
            )

        depot_s = {**DEPOT_START,    "latlon": depot_start_latlon}
        depot_e = {**depot_end_info, "latlon": depot_end_latlon}
        m       = build_map(ordered, letter, name, opt_dist,
                            day=day, depot_start=depot_s, depot_end=depot_e)

        results.append({
            "letter":           letter,
            "name":             name.replace("_", " "),
            "day":              day,
            "borough":          borough,
            "stops":            ordered,
            "orig_dist":        orig_dist,
            "opt_dist":         opt_dist,
            "box_counts":       box_counts,
            "map_html":         m._repr_html_() if m else None,
            "limit_warning":    limit_warning,
            "distance_warning": distance_warning,
            "depot_start":      DEPOT_START["label"],
            "depot_end":        depot_end_info["label"],
        })

        kitchen_rows.append({
            "Route":          f"Route {letter} — {name.replace('_', ' ')} ({day})",
            "Total Stops":    len(ordered),
            "Large":          box_counts.get("Large", 0),
            "Medium":         box_counts.get("Medium", 0),
            "Small":          box_counts.get("Small", 0),
            "Four-Date":      box_counts.get("Four-Date", 0),
            "Unknown":        box_counts.get("Unknown", 0),
            "Allergen Notes": "; ".join(allergen_notes),
        })

    return results, kitchen_rows, flags


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        password = request.form.get("password", "")
        expected = os.environ.get("APP_PASSWORD", "")
        if not expected or password == expected:
            session["authed"] = True
            return redirect(url_for("index"))
        return render_template("login.html", error="Incorrect password.")
    return render_template("login.html")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    if not _authed():
        return redirect(url_for("login"))
    cache     = load_cache()
    parse_id  = session.get("parse_id")
    parsed    = _store.get(parse_id, {}) if parse_id else {}
    error     = session.pop("error", None)
    return render_template(
        "index.html",
        cache_count=len(cache),
        parsed_count=parsed.get("count", 0),
        parse_flags=parsed.get("flags", []),
        distance_cap=MAX_ROUTE_MILES,
        contact_name="Oscar",
        error=error,
    )


@app.route("/upload", methods=["POST"])
def upload():
    if not _authed():
        return jsonify({"error": "Not authenticated"}), 401
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400
    stops, flags = parse_excel(file)
    if stops is None:
        return jsonify({"error": flags[0] if flags else "Failed to parse file"}), 400
    cache     = load_cache()
    new_count = sum(1 for s in stops if (s["addr1"], s["zipcode"]) not in cache)
    parse_id  = str(uuid.uuid4())
    _store[parse_id] = {"stops": stops, "flags": flags, "count": len(stops), "new_count": new_count}
    session["parse_id"] = parse_id
    return jsonify({"count": len(stops), "flags": flags, "new_count": new_count})


@app.route("/generate", methods=["POST"])
def generate():
    if not _authed():
        return redirect(url_for("login"))
    parse_id = session.get("parse_id")
    if not parse_id or parse_id not in _store:
        return redirect(url_for("index"))

    parsed       = _store[parse_id]
    distance_cap = float(request.form.get("distance_cap", MAX_ROUTE_MILES))
    contact_name = request.form.get("contact_name", "Oscar").strip() or "Oscar"

    try:
        results, kitchen_rows, flags = run_generation(
            parsed["stops"], parsed["flags"], distance_cap, contact_name
        )
    except RuntimeError as e:
        session["error"] = str(e)
        return redirect(url_for("index"))

    results_id = str(uuid.uuid4())
    _store[results_id] = {
        "results":        results,
        "kitchen_rows":   kitchen_rows,
        "flags":          flags,
        "generated_date": date.today().isoformat(),
        "contact_name":   contact_name,
    }
    session["results_id"] = results_id
    return redirect(url_for("results", results_id=results_id))


@app.route("/results/<results_id>")
def results(results_id):
    if not _authed():
        return redirect(url_for("login"))
    payload = _store.get(results_id)
    if not payload:
        session.pop("results_id", None)
        return redirect(url_for("index"))

    cache    = load_cache()
    parse_id = session.get("parse_id")
    parsed   = _store.get(parse_id, {}) if parse_id else {}

    r            = payload["results"]
    kitchen_rows = payload["kitchen_rows"]
    flags        = payload["flags"]
    gen_date     = payload["generated_date"]
    contact_name = payload["contact_name"]

    total_stops = sum(ro["Total Stops"] for ro in kitchen_rows)
    total_L     = sum(ro["Large"]       for ro in kitchen_rows)
    total_M     = sum(ro["Medium"]      for ro in kitchen_rows)
    total_S     = sum(ro["Small"]       for ro in kitchen_rows)
    total_FD    = sum(ro.get("Four-Date", 0) for ro in kitchen_rows)

    for route in r:
        route["household_clusters"] = detect_household_clusters(route["stops"])

    allergen_stops = [
        s for route in r for s in route["stops"]
        if s["allergens"] and s["allergens"].lower() not in ("none", "")
    ]

    return render_template(
        "results.html",
        results_id=results_id,
        results=r,
        kitchen_rows=kitchen_rows,
        flags=flags,
        gen_date=gen_date,
        contact_name=contact_name,
        total_stops=total_stops,
        total_L=total_L, total_M=total_M, total_S=total_S, total_FD=total_FD,
        n_flags=len(flags),
        allergen_stops=allergen_stops,
        MAX_STOPS_SOFT=MAX_STOPS_SOFT,
        cache_count=len(cache),
        parsed_count=parsed.get("count", 0),
        distance_cap=MAX_ROUTE_MILES,
    )


@app.route("/map/<results_id>/<letter>")
def route_map(results_id, letter):
    if not _authed():
        return "Unauthorized", 401
    payload = _store.get(results_id)
    if not payload:
        return "Not found", 404
    route = next((r for r in payload["results"] if r["letter"] == letter), None)
    if not route or not route.get("map_html"):
        return "Map unavailable", 404
    return route["map_html"], 200, {"Content-Type": "text/html"}


@app.route("/download/xlsx/<results_id>/<letter>")
def download_xlsx(results_id, letter):
    if not _authed():
        return redirect(url_for("login"))
    payload = _store.get(results_id)
    if not payload:
        return redirect(url_for("index"))
    route = next((r for r in payload["results"] if r["letter"] == letter), None)
    if not route:
        return "Not found", 404
    hh  = detect_household_clusters(route["stops"])
    buf = manifest_to_xlsx(route["stops"], route_info=route,
                           household_clusters=hh, contact=payload["contact_name"])
    fname = f"Route_{letter}_{route['name'].replace(' ', '_')}_manifest.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/download/csv/<results_id>/<letter>")
def download_csv(results_id, letter):
    if not _authed():
        return redirect(url_for("login"))
    payload = _store.get(results_id)
    if not payload:
        return redirect(url_for("index"))
    route = next((r for r in payload["results"] if r["letter"] == letter), None)
    if not route:
        return "Not found", 404
    data  = manifest_to_csv(route["stops"], route_info=route, contact=payload["contact_name"])
    fname = f"Route_{letter}_{route['name'].replace(' ', '_')}_manifest.csv"
    return send_file(io.BytesIO(data), mimetype="text/csv",
                     as_attachment=True, download_name=fname)


@app.route("/download/kitchen/<results_id>")
def download_kitchen(results_id):
    if not _authed():
        return redirect(url_for("login"))
    payload = _store.get(results_id)
    if not payload:
        return redirect(url_for("index"))
    data  = kitchen_to_csv(payload["kitchen_rows"])
    fname = f"Kitchen_Packing_List_{payload['generated_date']}.csv"
    return send_file(io.BytesIO(data), mimetype="text/csv",
                     as_attachment=True, download_name=fname)


@app.route("/clear-cache", methods=["POST"])
def clear_cache():
    if not _authed():
        return redirect(url_for("login"))
    try:
        CACHE_FILE.unlink(missing_ok=True)
    except OSError:
        pass
    return redirect(request.referrer or url_for("index"))


@app.route("/new", methods=["POST"])
def new_session():
    if not _authed():
        return redirect(url_for("login"))
    session.pop("results_id", None)
    session.pop("parse_id", None)
    return redirect(url_for("index"))


if __name__ == "__main__":
    app.run(debug=True)
