"""
app.py — Rethink Food Route Generator (Streamlit)
Run with: streamlit run app.py
"""

import csv
import io
import warnings
from datetime import date

import folium  # noqa: F401 — imported so folium map HTML renders correctly
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from geopy.geocoders import Nominatim

from rethink_routes import (
    BOX_COLORS,
    CACHE_FILE,
    DEPOT_BRONX_END,
    DEPOT_OTHER_END,
    DEPOT_START,
    MANIFEST_HEADERS,
    MAX_ROUTE_MILES,
    MAX_STOPS_HARD,
    MAX_STOPS_SOFT,
    ROUTES,
    ZIP_OVERRIDES,
    build_map,
    check_stop_limit,
    clean_box,
    detect_household_clusters,
    geocode_stop,
    load_cache,
    optimize_route,
    validate_geocode,
)

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Rethink Food — Route Generator",
    layout="wide",
)

# ── Global CSS ────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', -apple-system, sans-serif; }
h1 { font-weight: 800 !important; letter-spacing: -0.03em !important; }
h2 { font-weight: 700 !important; letter-spacing: -0.02em !important; }
h3 { font-weight: 600 !important; }

[data-testid="stSidebar"] {
    background-color: #111111 !important;
    border-right: 1px solid #242424;
}

[data-testid="stMetric"] {
    background-color: #171717;
    border: 1px solid #242424;
    border-radius: 10px;
    padding: 1rem 1.25rem !important;
}
[data-testid="stMetricLabel"] p {
    color: #777 !important;
    font-size: 0.72rem !important;
    text-transform: uppercase;
    letter-spacing: 0.08em;
}
[data-testid="stMetricValue"] { font-weight: 700 !important; }

.stTabs [data-baseweb="tab-list"] {
    background-color: #171717;
    border-radius: 8px;
    padding: 4px;
    gap: 2px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 6px;
    color: #777 !important;
    font-weight: 500;
}
.stTabs [aria-selected="true"] {
    background-color: #242424 !important;
    color: #f2f2f2 !important;
}
.stTabs [data-baseweb="tab-highlight"] { display: none; }

[data-testid="stExpander"] {
    border: 1px solid #242424 !important;
    border-radius: 8px !important;
}

hr { border-color: #242424 !important; }

.stDownloadButton > button {
    background-color: transparent !important;
    border: 1px solid #46F694 !important;
    color: #46F694 !important;
    font-weight: 600;
}
.stDownloadButton > button:hover {
    background-color: rgba(70, 246, 148, 0.08) !important;
}

[data-testid="stCaptionContainer"] p { color: #777 !important; }
</style>
""", unsafe_allow_html=True)

# ── Password gate ─────────────────────────────────────────────────────────────

def check_password():
    """Returns True if the user has entered the correct password."""
    # If no password is configured (e.g. local dev), skip the gate
    if "APP_PASSWORD" not in st.secrets:
        return True

    if st.session_state.get("authenticated"):
        return True

    with st.form("login"):
        st.markdown("""
<div style="text-align:center; padding:1rem 0 1.5rem;">
    <span style="display:inline-block; width:10px; height:10px; background:#46F694;
          border-radius:50%; margin-right:8px; vertical-align:middle;"></span>
    <span style="font-size:1.4rem; font-weight:800; letter-spacing:-0.02em;
          vertical-align:middle;">Rethink Food</span>
    <p style="color:#777; margin-top:0.5rem; font-size:0.9rem;">
        Route Generator &mdash; Team Access
    </p>
</div>
""", unsafe_allow_html=True)
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in", use_container_width=True)

    if submitted:
        if password == st.secrets["APP_PASSWORD"]:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")

    return False

if not check_password():
    st.stop()

# ── Session state defaults ────────────────────────────────────────────────────

for key, default in [
    ("results", None),
    ("kitchen_rows", []),
    ("flags", []),
    ("generated_date", None),
    ("parsed_stops", None),
    ("parsed_flags", []),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── Pure helpers ──────────────────────────────────────────────────────────────

def parse_excel(fileobj):
    """Read an uploaded Excel file and return (list[stop_dict], list[flag_str])."""
    wb = openpyxl.load_workbook(fileobj, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = rows[0]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    required = ["Member ID", "Box Size", "Address Line 1", "City", "State",
                "Zip", "Phone Number", "Delivery Instructions", "Status"]
    missing = [c for c in required if c not in col]
    if missing:
        return None, [f"Missing columns in spreadsheet: {missing}"]

    stops, flags = [], []
    for row in rows[1:]:
        status = str(row[col["Status"]] or "").strip().lower()
        if status != "active":
            continue

        zipcode = str(row[col["Zip"]] or "").replace(".0", "").strip().zfill(5)
        addr1   = str(row[col["Address Line 1"]] or "").strip()
        addr2   = str(row[col["Address Line 2"]] or "").strip() if "Address Line 2" in col else ""
        city    = str(row[col["City"]]   or "").strip()
        state   = str(row[col["State"]]  or "").strip()

        display_addr = addr1
        if addr2 and addr2.lower() not in ("none", ""):
            display_addr += f", {addr2}"
        display_addr += f", {city}, {state} {zipcode}"

        allergens = str(row[col["Meal Preferences/Allergens"]] or "").strip() \
            if "Meal Preferences/Allergens" in col else ""
        avail = str(row[col["Available Delivery Days"]] or "").strip() \
            if "Available Delivery Days" in col else ""
        notes_raw = " | ".join(filter(None, [
            str(row[col[c]] or "").strip()
            for c in ["Unnamed: 13", "Unnamed: 14", "Unnamed: 15"] if c in col
        ]))

        flag = ""
        if "cancel" in notes_raw.lower():
            flag = "Notes say 'cancel' but status is Active"
        elif "hold" in notes_raw.lower():
            flag = "Notes mention 'hold' — verify"
        if allergens and allergens.lower() not in ("none", ""):
            flag = (flag + " | " if flag else "") + f"Allergen: {allergens}"

        member_id = str(row[col["Member ID"]] or "").replace(".0", "")
        stop = {
            "member_id":             member_id,
            "addr1":                 addr1,
            "addr2":                 addr2,
            "city":                  city,
            "state":                 state,
            "zipcode":               zipcode,
            "display_addr":          display_addr,
            "phone":                 str(row[col["Phone Number"]] or "").strip(),
            "box_size":              clean_box(row[col["Box Size"]]),
            "allergens":             allergens,
            "delivery_instructions": str(row[col["Delivery Instructions"]] or "").strip(),
            "available_days":        avail,
            "notes":                 notes_raw,
            "flag":                  flag,
            "latlon":                None,
        }
        stops.append(stop)
        if flag:
            flags.append(f"Member {member_id} ({display_addr}): {flag}")

    return stops, flags


def _build_summary_lines(route_info, contact, total_members):
    """Return a list of summary-line strings for a route manifest header."""
    if "Bronx" in route_info.get("depot_end", ""):
        dep_end = DEPOT_BRONX_END
    else:
        dep_end = DEPOT_OTHER_END

    bc = route_info.get("box_counts", {})
    box_parts = [f"{sz}: {bc[sz]}" for sz in ["Large", "Medium", "Small", "Four-Date"] if bc.get(sz)]
    box_str = " | ".join(box_parts) if box_parts else "0"

    return [
        f"Route: Route {route_info['letter']} \u2014 {route_info['name']} ({route_info['day']})",
        f"Date: {date.today().isoformat()}",
        f"Start: {DEPOT_START['label']} ({DEPOT_START['addr1']} {DEPOT_START['zipcode']})",
        f"End: {dep_end['label']} ({dep_end['addr1']} {dep_end['zipcode']})",
        f"Point of Contact: {contact}",
        f"Total Members: {total_members}",
        f"Boxes: {box_str}",
    ]


def manifest_to_csv(ordered_stops, route_info=None, contact="Oscar"):
    """Build a CSV manifest with optional summary header rows."""
    buf = io.StringIO()
    w = csv.writer(buf)

    if route_info:
        for line in _build_summary_lines(route_info, contact, len(ordered_stops)):
            w.writerow([line])
        w.writerow(["---"])

    w.writerow(MANIFEST_HEADERS)
    for i, s in enumerate(ordered_stops, 1):
        w.writerow([
            i, s["member_id"], s["addr1"], s["addr2"], s["city"], s["zipcode"],
            s["phone"], s["box_size"], s["allergens"], s["delivery_instructions"],
            s["available_days"], s["notes"], s["flag"],
        ])
    return buf.getvalue().encode("utf-8")


def kitchen_to_csv(kitchen_rows):
    fields = ["Route", "Total Stops", "Large", "Medium", "Small", "Four-Date", "Unknown", "Allergen Notes"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fields)
    w.writeheader()
    w.writerows(kitchen_rows)
    if kitchen_rows:
        w.writerow({
            "Route": "TOTAL",
            "Total Stops": sum(r["Total Stops"] for r in kitchen_rows),
            "Large":       sum(r["Large"]       for r in kitchen_rows),
            "Medium":      sum(r["Medium"]      for r in kitchen_rows),
            "Small":       sum(r["Small"]       for r in kitchen_rows),
            "Four-Date":   sum(r.get("Four-Date", 0) for r in kitchen_rows),
            "Unknown":     sum(r["Unknown"]     for r in kitchen_rows),
            "Allergen Notes": "",
        })
    return buf.getvalue().encode("utf-8")


# ── XLSX formatted export ────────────────────────────────────────────────────

_XLSX_HEADERS = [
    "Stop", "Member ID", "Address", "Apt/Unit", "City", "Zip",
    "Phone", "Box Size", "Allergens", "Delivery Instructions",
    "Available Days", "Notes", "Flag", "Household",
]

_HOUSEHOLD_FILLS = [
    PatternFill(start_color="FFE8B0", end_color="FFE8B0", fill_type="solid"),  # light yellow
    PatternFill(start_color="B0D4FF", end_color="B0D4FF", fill_type="solid"),  # light blue
    PatternFill(start_color="FFB0C8", end_color="FFB0C8", fill_type="solid"),  # light pink
    PatternFill(start_color="B0FFB0", end_color="B0FFB0", fill_type="solid"),  # light green
    PatternFill(start_color="D4B0FF", end_color="D4B0FF", fill_type="solid"),  # light lavender
    PatternFill(start_color="FFD4B0", end_color="FFD4B0", fill_type="solid"),  # light peach
]


def manifest_to_xlsx(ordered_stops, route_info, household_clusters, contact="Oscar"):
    """Create a styled .xlsx manifest and return as io.BytesIO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"

    num_cols = len(_XLSX_HEADERS)

    # -- Summary section (rows 1-8) -------------------------------------------
    summary = _build_summary_lines(route_info, contact, len(ordered_stops))

    # Row 1: Route name (bold, 14pt, merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=summary[0].replace("Route: ", ""))
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="left")

    # Rows 2-7: remaining summary lines
    for r_idx, line in enumerate(summary[1:], start=2):
        ws.cell(row=r_idx, column=1, value=line)

    # Row 8: blank separator (no content)

    # -- Column headers (row 9) -----------------------------------------------
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    HEADER_ROW = 9

    for c_idx, hdr in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c_idx, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # -- Data rows (row 10+) --------------------------------------------------
    alt_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    red_font = Font(color="FF0000")

    # Map household group letters to fill indices
    group_letters = sorted(set(household_clusters.values()))
    group_fill_map = {
        letter: _HOUSEHOLD_FILLS[i % len(_HOUSEHOLD_FILLS)]
        for i, letter in enumerate(group_letters)
    }

    # Track max widths for auto-fit (start with header lengths)
    col_widths = [len(h) for h in _XLSX_HEADERS]

    DATA_START = HEADER_ROW + 1
    for i, s in enumerate(ordered_stops):
        row_num = DATA_START + i
        hh_letter = household_clusters.get(i, "")
        values = [
            i + 1,
            s["member_id"],
            s["addr1"],
            s["addr2"],
            s["city"],
            s["zipcode"],
            s["phone"],
            s["box_size"],
            s["allergens"],
            s["delivery_instructions"],
            s["available_days"],
            s["notes"],
            s["flag"],
            hh_letter,
        ]

        is_flagged = bool(s["flag"])
        is_alt_row = i % 2 == 1
        hh_fill = group_fill_map.get(hh_letter)

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)

            # Flagged stops: red font
            if is_flagged:
                cell.font = red_font

            # Household cluster fill takes priority, then alternating gray
            if hh_fill:
                cell.fill = hh_fill
            elif is_alt_row:
                cell.fill = alt_fill

            # Update column width tracker
            val_len = len(str(val)) if val else 0
            if val_len > col_widths[c_idx - 1]:
                col_widths[c_idx - 1] = val_len

    # -- Auto-fit column widths -----------------------------------------------
    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = min(w + 2, 45)

    # -- Write to bytes -------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Route generation (runs on button click) ───────────────────────────────────

def run_generation(all_stops, all_flags):
    geocache   = load_cache()
    geolocator = Nominatim(user_agent="rethink_food_app")

    # Build zip -> routes mapping (supports overlapping zips across days)
    zip_to_routes = {}
    for letter, name, borough, day, zips in ROUTES:
        for z in zips:
            zip_to_routes.setdefault(z, []).append((letter, name, borough, day))

    route_buckets = {letter: [] for letter, *_ in ROUTES}
    flags = list(all_flags)

    for stop in all_stops:
        # Check zip-to-route overrides first
        override_letter = ZIP_OVERRIDES.get(stop["zipcode"])
        if override_letter and override_letter in route_buckets:
            # Find the borough for the override route
            override_borough = next(
                (borough for letter, name, borough, day, zips in ROUTES if letter == override_letter),
                "Queens"
            )
            route_buckets[override_letter].append((stop, override_borough))
            continue

        matching = zip_to_routes.get(stop["zipcode"], [])
        if not matching:
            flags.append(
                f"Member {stop['member_id']} (zip {stop['zipcode']}) "
                "not assigned to any route — add zip to ROUTES in rethink_routes.py"
            )
        elif len(matching) == 1:
            route_buckets[matching[0][0]].append((stop, matching[0][2]))
        else:
            # Load-balance: assign to the matching route with fewest stops so far
            best = min(matching, key=lambda r: len(route_buckets[r[0]]))
            route_buckets[best[0]].append((stop, best[2]))

    # ── Geocode depots ────────────────────────────────────────────────────────
    depot_start_latlon = geocode_stop(
        geolocator, DEPOT_START["addr1"], DEPOT_START["zipcode"], DEPOT_START["borough"], geocache)
    depot_bronx_latlon = geocode_stop(
        geolocator, DEPOT_BRONX_END["addr1"], DEPOT_BRONX_END["zipcode"], DEPOT_BRONX_END["borough"], geocache)
    depot_other_latlon = geocode_stop(
        geolocator, DEPOT_OTHER_END["addr1"], DEPOT_OTHER_END["zipcode"], DEPOT_OTHER_END["borough"], geocache)

    if not depot_start_latlon or not depot_bronx_latlon or not depot_other_latlon:
        st.error(
            "Failed to geocode one or more depot addresses. "
            "Check your internet connection and try again."
        )
        return

    # ── Geocoding progress ────────────────────────────────────────────────────
    n_cached = sum(1 for s in all_stops if (s["addr1"], s["zipcode"]) in geocache)
    n_new    = len(all_stops) - n_cached

    geo_header = st.empty()
    geo_header.info(
        f"**{n_cached}** addresses loaded from cache &nbsp;|&nbsp; "
        f"**{n_new}** new address(es) to geocode"
        + (f" (~{n_new}s)" if n_new else "")
    )

    progress_bar = st.progress(0, text="Starting geocoding...")
    status_text  = st.empty()

    for i, stop in enumerate(all_stops):
        matching = zip_to_routes.get(stop["zipcode"])
        borough  = matching[0][2] if matching else "Queens"

        cached = (stop["addr1"], stop["zipcode"]) in geocache
        if not cached:
            status_text.caption(f"Geocoding: {stop['addr1']}, {stop['zipcode']}...")

        result         = geocode_stop(geolocator, stop["addr1"], stop["zipcode"], borough, geocache)
        stop["latlon"] = result

        if result is None and not cached:
            flags.append(
                f"Member {stop['member_id']} ({stop['addr1']}, {stop['zipcode']}): "
                "geocoding failed — stop excluded from map"
            )

        progress_bar.progress((i + 1) / len(all_stops),
                              text=f"Geocoding {i + 1} / {len(all_stops)}")

    progress_bar.progress(1.0, text="Geocoding complete.")
    status_text.empty()

    # ── Per-route optimization ────────────────────────────────────────────────
    results, kitchen_rows = [], []

    with st.spinner("Optimizing routes..."):
        for letter, name, borough, day, zips in ROUTES:
            entries = route_buckets[letter]
            if not entries:
                continue

            geocoded = [e[0] for e in entries if e[0]["latlon"]]
            if not geocoded:
                continue

            depot_end_latlon = depot_bronx_latlon if borough == "Bronx" else depot_other_latlon
            depot_end_info   = DEPOT_BRONX_END    if borough == "Bronx" else DEPOT_OTHER_END

            ordered, orig_dist, opt_dist = optimize_route(
                geocoded, depot_start_latlon, depot_end_latlon)

            # Stop limit check
            limit_warning = check_stop_limit(ordered)
            if limit_warning:
                flags.append(f"Route {letter} ({name.replace('_',' ')}): {limit_warning}")

            # Distance cap check
            dist_cap = st.session_state.get("distance_cap", MAX_ROUTE_MILES)
            distance_warning = None
            if opt_dist > dist_cap:
                distance_warning = f"{opt_dist:.1f} miles exceeds {dist_cap:.1f} mile cap"
                flags.append(f"Route {letter} ({name.replace('_',' ')}): {distance_warning}")

            box_counts     = {"Large": 0, "Medium": 0, "Small": 0, "Four-Date": 0, "Unknown": 0}
            allergen_notes = []
            for s in ordered:
                box_counts[s["box_size"]] = box_counts.get(s["box_size"], 0) + 1
                if s["allergens"] and s["allergens"].lower() not in ("none", ""):
                    allergen_notes.append(f"Member {s['member_id']}: {s['allergens']}")

            # Validate box counts match stop count
            if sum(box_counts.values()) != len(ordered):
                flags.append(
                    f"Route {letter} ({name.replace('_',' ')}): box count mismatch — "
                    f"{sum(box_counts.values())} boxes counted vs {len(ordered)} stops"
                )

            depot_s  = {**DEPOT_START,    "latlon": depot_start_latlon}
            depot_e  = {**depot_end_info, "latlon": depot_end_latlon}
            m        = build_map(ordered, letter, name, opt_dist,
                                 day=day, depot_start=depot_s, depot_end=depot_e)
            map_html = m._repr_html_() if m else None

            results.append({
                "letter":           letter,
                "name":             name.replace("_", " "),
                "day":              day,
                "borough":          borough,
                "stops":            ordered,
                "orig_dist":        orig_dist,
                "opt_dist":         opt_dist,
                "box_counts":       box_counts,
                "map_html":         map_html,
                "limit_warning":    limit_warning,
                "distance_warning": distance_warning,
                "depot_start":      DEPOT_START["label"],
                "depot_end":        depot_end_info["label"],
            })

            kitchen_rows.append({
                "Route":          f"Route {letter} — {name.replace('_', ' ')} ({day})",
                "Total Stops":    len(ordered),
                "Large":          box_counts.get("Large", 0),
                "Medium":         box_counts.get("Medium", 0),
                "Small":          box_counts.get("Small", 0),
                "Four-Date":      box_counts.get("Four-Date", 0),
                "Unknown":        box_counts.get("Unknown", 0),
                "Allergen Notes": "; ".join(allergen_notes),
            })

    st.session_state.results        = results
    st.session_state.kitchen_rows   = kitchen_rows
    st.session_state.flags          = flags
    st.session_state.generated_date = date.today().isoformat()


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
<div style="margin-bottom:0.75rem;">
    <div style="display:flex; align-items:center; gap:8px; margin-bottom:2px;">
        <div style="width:10px; height:10px; background:#46F694; border-radius:50%;
             flex-shrink:0;"></div>
        <span style="font-weight:800; font-size:1.05rem; letter-spacing:-0.01em;">
            Rethink Food
        </span>
    </div>
    <span style="color:#666; font-size:0.72rem; text-transform:uppercase;
          letter-spacing:0.1em; padding-left:18px;">Route Generator</span>
</div>
""", unsafe_allow_html=True)
    st.divider()

    uploaded_file = st.file_uploader(
        "Upload member list (.xlsx)",
        type=["xlsx"],
        help="Upload the weekly MTM members Excel spreadsheet.",
    )

    if uploaded_file is not None:
        with st.spinner("Reading file..."):
            stops, parse_flags = parse_excel(uploaded_file)

        if stops is None:
            st.error(parse_flags[0] if parse_flags else "Failed to parse file.")
        else:
            st.session_state.parsed_stops = stops
            st.session_state.parsed_flags = parse_flags
            st.success(f"{len(stops)} active members found")

    st.divider()

    distance_cap = st.number_input(
        "Max route distance (miles)",
        value=MAX_ROUTE_MILES,
        min_value=10.0,
        max_value=60.0,
        step=1.0,
        help="Routes exceeding this mileage will be flagged.",
    )
    st.session_state["distance_cap"] = distance_cap

    contact_name = st.text_input(
        "Point of Contact",
        value="Oscar",
        help="Name printed on each route manifest.",
    )
    st.session_state["contact_name"] = contact_name

    st.divider()

    can_generate = bool(st.session_state.get("parsed_stops"))
    generate_btn = st.button(
        "Generate Routes",
        type="primary",
        disabled=not can_generate,
        use_container_width=True,
    )

    st.divider()

    cache = load_cache()
    st.caption(f"Geocode cache: **{len(cache)}** saved addresses")
    st.caption("Cached addresses are never re-geocoded, saving time each week.")

    if st.button("Clear geocode cache", use_container_width=True):
        CACHE_FILE.unlink(missing_ok=True)
        st.toast("Cache cleared.")
        st.rerun()

# ── Main area ─────────────────────────────────────────────────────────────────

if generate_btn and st.session_state.get("parsed_stops"):
    st.title("Generating routes...")
    run_generation(
        st.session_state.parsed_stops,
        st.session_state.parsed_flags,
    )
    st.rerun()

results = st.session_state.get("results")

# ── Empty / landing state ─────────────────────────────────────────────────────

if not results:
    st.markdown("""
<div style="margin-bottom:2rem;">
    <h1 style="font-size:2.2rem; font-weight:800; letter-spacing:-0.04em; margin-bottom:0.5rem;">
        Route Generator
    </h1>
    <p style="color:#777; font-size:1rem; margin:0;">
        Upload a member list in the sidebar, then click
        <span style="color:#46F694; font-weight:600;">Generate Routes</span>.
    </p>
</div>
<div style="display:grid; grid-template-columns:1fr 1fr; gap:1rem;">
    <div style="background:#171717; border:1px solid #242424; border-radius:12px; padding:1.5rem;">
        <div style="color:#46F694; font-size:0.68rem; text-transform:uppercase;
             letter-spacing:0.12em; margin-bottom:1rem; font-weight:700;">What you'll get</div>
        <ul style="color:#ccc; margin:0; padding-left:1.1rem; line-height:2; font-size:0.9rem;">
            <li>Optimized interactive map for each driver route</li>
            <li>Downloadable stop manifest per route</li>
            <li>Kitchen packing list (box counts + allergens)</li>
            <li>Flags report for anomalies to review</li>
        </ul>
    </div>
    <div style="background:#171717; border:1px solid #242424; border-radius:12px; padding:1.5rem;">
        <div style="color:#46F694; font-size:0.68rem; text-transform:uppercase;
             letter-spacing:0.12em; margin-bottom:1rem; font-weight:700;">How it works</div>
        <ul style="color:#ccc; margin:0; padding-left:1.1rem; line-height:2; font-size:0.9rem;">
            <li>Members assigned to 14 routes across Mon&ndash;Fri</li>
            <li>Stops sequenced from Tribeca depot using TSP optimization</li>
            <li>Addresses cached &mdash; subsequent weeks are instant</li>
            <li>Routes over 25 stops (28 if dense) flagged automatically</li>
        </ul>
    </div>
</div>
""", unsafe_allow_html=True)
    st.stop()

# ── Results ───────────────────────────────────────────────────────────────────

gen_date     = st.session_state.get("generated_date", date.today().isoformat())
kitchen_rows = st.session_state.get("kitchen_rows", [])
flags        = st.session_state.get("flags", [])

st.title(f"Routes — Week of {gen_date}")

# Top-line metrics
total_stops = sum(r["Total Stops"] for r in kitchen_rows)
total_L     = sum(r["Large"]       for r in kitchen_rows)
total_M     = sum(r["Medium"]      for r in kitchen_rows)
total_S     = sum(r["Small"]       for r in kitchen_rows)
total_FD    = sum(r.get("Four-Date", 0) for r in kitchen_rows)
n_flags     = len(flags)

m1, m2, m3, m4, m5, m6 = st.columns(6)
m1.metric("Total Stops",    total_stops)
m2.metric("Large Boxes",    total_L)
m3.metric("Medium Boxes",   total_M)
m4.metric("Small Boxes",    total_S)
m5.metric("Four-Date Boxes", total_FD)
m6.metric("Flags",          n_flags,
          delta=f"{n_flags} to review" if n_flags else "All clear",
          delta_color="inverse" if n_flags else "off")

st.divider()

# Tabs: one per route + Kitchen + Flags
tab_labels = [f"Route {r['letter']}" for r in results] + ["Kitchen List", "Flags"]
tabs       = st.tabs(tab_labels)

for i, route in enumerate(results):
    with tabs[i]:
        # Header row
        left, right = st.columns([4, 1])
        with left:
            st.subheader(f"Route {route['letter']} — {route['name']}")
            saved_pct = ((route['orig_dist'] - route['opt_dist']) / route['orig_dist'] * 100) \
                if route['orig_dist'] > 0 else 0
            st.caption(
                f"{route['day']}  ·  "
                f"{len(route['stops'])} stops  ·  "
                f"**{route['opt_dist']:.1f} mi** total (incl. depot legs)  ·  "
                f"saved {saved_pct:.0f}%"
            )
            st.caption(
                f"Start: {route['depot_start']}  →  End: {route['depot_end']}"
            )
        with right:
            contact = st.session_state.get("contact_name", "Oscar")
            hh_clusters = detect_household_clusters(route["stops"])
            xlsx_bytes = manifest_to_xlsx(
                route["stops"], route_info=route,
                household_clusters=hh_clusters, contact=contact,
            )
            st.download_button(
                label="Download Manifest (.xlsx)",
                data=xlsx_bytes,
                file_name=f"Route_{route['letter']}_{route['name'].replace(' ','_')}_manifest.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.download_button(
                label="Download CSV",
                data=manifest_to_csv(route["stops"], route_info=route, contact=contact),
                file_name=f"Route_{route['letter']}_{route['name'].replace(' ','_')}_manifest.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # Stop limit warning
        if route.get("limit_warning"):
            n = len(route["stops"])
            if n > MAX_STOPS_SOFT:
                st.error(f"Over capacity: {route['limit_warning']}")
            else:
                st.warning(f"Stop limit: {route['limit_warning']}")

        # Distance cap warning
        if route.get("distance_warning"):
            st.warning(f"Distance cap: {route['distance_warning']}")

        # Map
        if route["map_html"]:
            components.html(route["map_html"], height=530, scrolling=False)
        else:
            st.warning("Map unavailable — all stops in this route failed geocoding.")

        # Box summary
        bc = route["box_counts"]
        b1, b2, b3, b4, b5 = st.columns(5)
        b1.metric("Large",     bc.get("Large", 0))
        b2.metric("Medium",    bc.get("Medium", 0))
        b3.metric("Small",     bc.get("Small", 0))
        b4.metric("Four-Date", bc.get("Four-Date", 0))
        b5.metric("Flagged",   sum(1 for s in route["stops"] if s["flag"]))

        # Manifest table (collapsible)
        with st.expander("View full stop manifest"):
            manifest_data = [
                {
                    "Stop":         j,
                    "Member ID":    s["member_id"],
                    "Address":      s["display_addr"],
                    "Phone":        s["phone"],
                    "Box":          s["box_size"],
                    "Instructions": s["delivery_instructions"],
                    "Avail. Days":  s["available_days"],
                    "Notes":        s["notes"],
                    "Flag":         s["flag"],
                    "Household":    hh_clusters.get(j - 1, ""),
                }
                for j, s in enumerate(route["stops"], 1)
            ]
            st.dataframe(
                pd.DataFrame(manifest_data),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Stop":      st.column_config.NumberColumn(width="small"),
                    "Box":       st.column_config.TextColumn(width="small"),
                    "Flag":      st.column_config.TextColumn(width="medium"),
                    "Household": st.column_config.TextColumn(width="small"),
                },
            )

# Kitchen packing list tab
with tabs[len(results)]:
    st.subheader("Kitchen Packing List")
    st.caption(
        f"Total: **{total_stops} deliveries** — "
        f"{total_L} Large / {total_M} Medium / {total_S} Small / {total_FD} Four-Date"
    )

    kitchen_df = pd.DataFrame(kitchen_rows)
    totals_row = pd.DataFrame([{
        "Route":         "**TOTAL**",
        "Total Stops":   total_stops,
        "Large":         total_L,
        "Medium":        total_M,
        "Small":         total_S,
        "Four-Date":     total_FD,
        "Unknown":       sum(r.get("Unknown", 0) for r in kitchen_rows),
        "Allergen Notes": "",
    }])
    st.dataframe(
        pd.concat([kitchen_df, totals_row], ignore_index=True),
        use_container_width=True,
        hide_index=True,
    )

    st.download_button(
        label="Download Kitchen Packing List",
        data=kitchen_to_csv(kitchen_rows),
        file_name=f"Kitchen_Packing_List_{gen_date}.csv",
        mime="text/csv",
    )

    # Allergen callouts
    allergen_stops = [
        s for r in results for s in r["stops"]
        if s["allergens"] and s["allergens"].lower() not in ("none", "")
    ]
    if allergen_stops:
        st.divider()
        st.markdown("#### Allergen / Dietary Notes")
        for s in allergen_stops:
            st.warning(
                f"Member {s['member_id']} — {s['addr1']}: **{s['allergens']}**"
            )
    else:
        st.info("No dietary restrictions or allergens flagged for this week.")

# Flags tab
with tabs[len(results) + 1]:
    st.subheader("Flags — Review Before Loading Trucks")
    if flags:
        st.warning(f"{len(flags)} item(s) require attention.")
        for f in flags:
            st.markdown(f"- {f}")
    else:
        st.success("No flags — all clear for delivery.")
